import sys
import openpyxl

total_population = 5607000
#peopl
total_landArea = 278.2
#mile^2
total_gdp = 297000000000
#billion $ USD 2016

def populationDensityCalc():
    return total_population/total_landArea

def gdpPercapita():
    return total_gdp/total_population

def calcWorkingPopLocal():
    wb1 = openpyxl.load_workbook(
        'data/workBook2.xlsx')  # load workBook2.xlsx this is about the household that is working etc.
    sheet1 = wb1['Title']
    # total_amount_of_households = sheet1.cell(row=8, column=4).value # stop at 21
    # total_amount_of_households_thats_not_working = sheet1.cell(row=7, column=4) # stop at 21
    for x in range(4, 22):
        percentage_of_household_not_working = sheet1.cell(row=7, column=x).value/sheet1.cell(row=8,column=x).value
        year = sheet1.cell(row=6, column=x).value
        printf("Percentage of household that is not working from year: %s,", year)
        printf("Percentage of household that is not working is: %f \n", percentage_of_household_not_working)

def laborForceCalc():
    wb1 = openpyxl.load_workbook(
        'data/labourForceAged15_over.xlsx')  # labourForceAged15_over.xlsx about employment rate
    sheet1 = wb1['Title']
    # first compare the total workforce pop and the unemployment rate of local residence pr + citizen
    for x in range(2,20): # 2 ~ 19 from 2000 to 2017
        year = sheet1.cell(row=6,column=x).value
        total_laborForce = sheet1.cell(row=7,column=x).value
        unemployeed_percentage = sheet1.cell(row=9,column=x).value/total_laborForce
        printf("The year is: %s, unemployed percentage is:| %f | base total employeable people | %f thousands|\n", year, unemployeed_percentage, total_laborForce)

def compareEmployementLocalVSmigrant():
    wb1 = openpyxl.load_workbook(
        'data/labourForceAged15_over.xlsx')  # labourForceAged15_over.xlsx about employment rate
    sheet1 = wb1['Title']
    print("---Local - total---")
    for x in range(2,20): # 2 ~ 19 from 2000 to 2017
        year = sheet1.cell(row=6,column=x).value
        total_unemployeed = sheet1.cell(row=10,column=x).value
        unemployeed_differences = sheet1.cell(row=11,column=x).value - total_unemployeed
        printf("The year is: %s, unemployed difference is:| %f | residence unemployment rate is: | %f | total is: | %f | \n"
               , year, unemployeed_differences, sheet1.cell(row=11,column=x).value, total_unemployeed)

def calcOfworkVISA():
    wb1 = openpyxl.load_workbook(
        'data/foreign-workforce-numbers.xlsx')  # visa//work visa
    sheet1 = wb1['Sheet1']
    wb2 = openpyxl.load_workbook(
        'data/labourForceAged15_over.xlsx')  # labourForceAged15_over.xlsx about employment rate
    sheet2 = wb2['Title']
    foreign_workers = sheet1.cell(row=9,column=2).value
    total_laborForce = [sheet2.cell(row=7,column=14).value * 1000, sheet2.cell(row=7,column=15).value * 1000,
                        sheet2.cell(row=7,column=16).value * 1000,
                        sheet2.cell(row=7,column=17).value * 1000, sheet2.cell(row=7,column=18).value * 1000,
                        sheet2.cell(row=7,column=19).value * 1000]
    for x in range(2,8): # 2 - 8 from 2012 - 2017
        year = sheet1.cell(row=2,column=x).value
        result_percent = sheet1.cell(row=9, column=x).value/total_laborForce[x-2]
        printf("The percentage of foreign work force in total work force is: | %f | and time is: %s \n", result_percent, year)
    for x in range(2, 8):
        year = sheet1.cell(row=2, column=x).value
        result_percent = (sheet1.cell(row=9, column=x).value-sheet1.cell(row=10, column=x).value)/sheet1.cell(row=9, column=x).value
        printf("The percentage of construction and domestic worker pop in foreign work force is: | %f | and time is: %s \n", result_percent,year)

def calcConstructWorkers():
    wb1 = openpyxl.load_workbook(
        'data/foreign-workforce-numbers.xlsx')  # visa//work visa
    sheet1 = wb1['Sheet1']
    for x in range(2, 8):  # 2 - 8 from 2012 - 2017
        year = sheet1.cell(row=2, column=x).value
        result_percent = (sheet1.cell(row=10, column=x).value - sheet1.cell(row=11, column=x).value)
        printf("The amount of construction worker is: | %f | and time is: %s \n", result_percent,year)


def printf(format, *args):
    sys.stdout.write(format % args)



def main():
    printf("Population Density is: %d people/mi^2 \n", populationDensityCalc())
    printf("GDP per capita is: %d USD \n", gdpPercapita())
    calcWorkingPopLocal()
    laborForceCalc()
    compareEmployementLocalVSmigrant()
    calcOfworkVISA()
    calcConstructWorkers()


main()

