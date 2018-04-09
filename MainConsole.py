# Haomin Shi City Study - Singapore - for SSCI 204
# 04/09/2018

# imports
import openpyxl
import sys

############

wb1 = openpyxl.load_workbook('data/workBook1.xlsx')  # load a1.xls A1- Resident Population by Age Group, Ethnic Group, Sex and Residential Status.xls
sheet1 = wb1['Sheet1']

# processing Sheets
# basic percentage of population
def process1(sheetT):

    if sheetT == sheet1:
        print(type(sheetT))
        totalPopulation = sheetT.cell(row=3,column=2).value
        male_workforce_total = 0
        female_workforce_total = 0
        maleTotalPopulation = sheetT.cell(row=3,column=3).value
        femaleTotalPopulation = sheetT.cell(row=3,column=4).value

        print("!!!MALE!!! data below")
        for x in range(4, 22):
            percentage = float((sheetT.cell(row=x, column=3).value)/totalPopulation)
            printf('age group is: %s, percentage of male in total population is: %f \n', sheetT.cell(row=x, column=1).value, percentage)
        print("-----------------------------------------------------------------")
        print("!!!FEMALE!!! data below")
        for x in range(4, 22):
            percentage = float((sheetT.cell(row=x, column=4).value)/totalPopulation)
            printf('age group is: %s, percentage of female in total population is: %f \n', sheetT.cell(row=x, column=1).value, percentage)

        # in singapore the age for retirement is 62
        # thus the male labor pool is roughly this
        for x in range(7,22-6): # x need to start form 7, ends at 22-6
            male_workforce_total += sheetT.cell(row=x, column=3).value
        for x in range(7,22-6): # x need to start form 7, ends at 22-6
            female_workforce_total += sheetT.cell(row=x, column=4).value

        total_male_workforce_percentage = male_workforce_total/totalPopulation
        total_female_workforce_percentage = female_workforce_total / totalPopulation

        total_male_workforce_percentage_inM = male_workforce_total / maleTotalPopulation
        total_female_workforce_percentage_inF = female_workforce_total / femaleTotalPopulation

        print("-----------------------------------------------------------------")
        printf("Male workforce total's percentage in total population is: %f \n", total_male_workforce_percentage)
        print("-----------------------------------------------------------------")
        printf("Female workforce total's percentage in total population is: %f\n", total_female_workforce_percentage)
        print("-----------------------------------------------------------------")
        printf("Together they are: %f \n", total_male_workforce_percentage + total_male_workforce_percentage)
        print("-----------------------------------------------------------------")
        printf("based on the data, the percentage of female that chooses to work is %f \n", total_female_workforce_percentage_inF)
        printf("based on the data, the percentage of male that chooses to work is %f \n", total_male_workforce_percentage_inM)

        femaleOverTotalPop = sheetT.cell(row=3, column=4).value / totalPopulation
        maleOverTotalPop = 1-femaleOverTotalPop
        printf("just for the information: female percentage in total population is: %f And for male is: %f \n", femaleOverTotalPop, maleOverTotalPop)

        # some interesting comparison might be find the workfoce percentage of foreigner pop in singapore and compare with the locals
        # i will use the PR population to compare with the Citizen first
        # some interesting finding, the tot
        PR_male_workforce_total = 0
        PR_female_workforce_total = 0
        PR_totalPopulation = sheetT.cell(row=47, column=2).value
        PR_maleTotalPopulation = sheetT.cell(row=47, column=3).value
        PR_femaleTotalPopulation = sheetT.cell(row=47, column=4).value
        for x in range(52,65-6): # x need to start form 7, ends at 22-6
            PR_male_workforce_total += sheetT.cell(row=x, column=3).value
        for x in range(52,65-6): # x need to start form 7, ends at 22-6
            PR_female_workforce_total += sheetT.cell(row=x, column=4).value
        PR_total_male_workforce_percentage = PR_male_workforce_total / PR_totalPopulation
        PR_total_female_workforce_percentage = PR_female_workforce_total / PR_totalPopulation
        print("-----------------------------------------------------------------")
        printf("PR Male workforce total's percentage in PR total population is: %f \n", PR_total_male_workforce_percentage)
        print("-----------------------------------------------------------------")
        printf("PR Female workforce total's percentage in PR total population is: %f\n", PR_total_female_workforce_percentage)
        print("-----------------------------------------------------------------")
        printf("Together they are: %f \n", PR_total_male_workforce_percentage + PR_total_male_workforce_percentage)
        print("-----------------------------------------------------------------")
        # in the case of PR, the male to choose to work, and the female that chooses to work
        PR_total_male_workforce_percentage_inM = PR_male_workforce_total / PR_maleTotalPopulation
        PR_total_female_workforce_percentage_inF = PR_female_workforce_total / PR_femaleTotalPopulation
        printf("based on the data, the percentage of female PR that chooses to work is %f \n", PR_total_female_workforce_percentage_inF)
        printf("based on the data, the percentage of male PR that chooses to work is %f \n", PR_total_male_workforce_percentage_inM)
        # got some interesting finding, the PR work percentage is much higher 5% ish.
        # and there are more female PR choose to work 2%
        print("-----------------------------------------------------------------")
        PR_femaleOverTotalPR = sheetT.cell(row=47, column=4).value / PR_totalPopulation
        PR_maleOverTotalPR = 1-PR_femaleOverTotalPR
        printf("just for the information: PR female percentage in total PR population is: %f And for PR male is: %f \n",PR_femaleOverTotalPR, PR_maleOverTotalPR)
        print("-----------------------------------------------------------------")

###########################################################################
# print helper
def printf(format, *args):
    sys.stdout.write(format % args)

def main(): # main console
    print("Welcome to Console\n")
    print("in work book 1 you have: ")

    print(*wb1, sep=", ")
    while True:
        print("type in 1 to start")
        print("type in exit to quit")
        usrIn = input(">>>Type in Command $ ")
        if usrIn == "exit":
            break
        if usrIn == "1":
            process1(sheet1)


##########################################################################
main()